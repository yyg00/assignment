import cv2
import numpy as np
import pytesseract
import sys
import openpyxl


def process_img(path):
    img = cv2.imread(path, 1)
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    # data augmentation => high resolution
    blur = cv2.GaussianBlur(gray, (3,3), 0) 
    thresh = cv2.adaptiveThreshold(~blur, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 35, -5)
    denoised = cv2.medianBlur(thresh, 3)
    rows, cols = denoised.shape
    precision = 10 
    kernel  = cv2.getStructuringElement(cv2.MORPH_RECT, (cols // precision, 1))
    eroded = cv2.erode(denoised, kernel, iterations = 1)
    row_lines = cv2.dilate(eroded, kernel, iterations = 1)
    kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, rows // precision))
    eroded = cv2.erode(denoised, kernel, iterations = 1)
    col_lines = cv2.dilate(eroded, kernel, iterations = 1)
    intersections = cv2.bitwise_and(row_lines, col_lines)
    bitwise_and = cv2.bitwise_and(col_lines, row_lines)
    xs,ys = np.where(bitwise_and > 0)
    x_list, y_list=[], []
    min_dist = 5
    xs = np.sort(xs)
    for i in range(len(xs)-1):
        if(xs[i+1]-xs[i] > min_dist):
            x_list.append(xs[i])
    x_list.append(xs[i])
    ys = np.sort(ys)
    for i in range(len(ys)-1):
        if(ys[i+1]-ys[i] > min_dist):
            y_list.append(ys[i])
    y_list.append(ys[i])
    form_line = cv2.add(row_lines, col_lines)
    contours, hierarchy = cv2.findContours(form_line, cv2.RETR_TREE, cv2.CHAIN_APPROX_NONE)
    contours = list(contours)
    contours.reverse()
    image_copy = gray.copy()
    msg = []
    bx_idxs = []
    for i in range(0,len(contours)-1):
        x, y, w, h = cv2.boundingRect(contours[i])
        im = image_copy[y: y + h, x: x + w]
        bx_idx = []
        for pt in x_list:
            if y in range(pt-min_dist,pt+min_dist+1):
                bx_idx.append(pt)
                break
        for pt in y_list:
            if x in range(pt-min_dist, pt+min_dist+1):
                bx_idx.append(pt)
                break
        bx_idx.append(h)
        bx_idx.append(w)
        bx_idxs.append(bx_idx)
        text = pytesseract.image_to_string(im, lang='eng')
        text = text.replace('\n', ' ')
        msg.append(text.strip())
    zipped = zip(bx_idxs, msg)
    s = sorted(zipped, key=lambda x: (x[0][0], x[0][1]))
    boxes, msg = zip(*s)
    return boxes, msg

def output_to_excel(out_file_name, boxes, msg):
    wb = openpyxl.Workbook()
    ws = wb.active
    for box, text in zip(boxes, msg):
        y, x, h, w = box
        col = x // 10 + 1
        row = y // 10 + 1
        end_col = (x + w) // 10
        end_row = (y + h) // 10
        ws.merge_cells(start_row=row, start_column=col, end_row=end_row, end_column=end_col)
        for i in range(col, end_col+1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w / 10
        for i in range(row, end_row+1):
            ws.row_dimensions[i].height = h

        ws.cell(row=row, column=col).value = str(text)
    wb.save(out_file_name)
    wb.close()

if __name__ == '__main__':
    image, excel = sys.argv[1], sys.argv[2]
    box, msg = process_img(image)
    output_to_excel(excel, box, msg)